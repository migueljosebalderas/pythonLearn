"""
1. Verificar archivos xlsx del directorio indicado
2. Crear archivo resultado.xlsx
3. Por cada archivo
    Identificar el origen (IBM, TERA)
    Leer cada una de las hojas
        Leer cada registro
            Agregar $ORIGEN
            Guardar los datos del registro una hoja (en un archivo result.xlsx)
4. Grabar archivo

"""


import sys
from openpyxl import Workbook
from openpyxl import load_workbook
from os import listdir

wb_result = Workbook()
nueva = True

data_dir = ".\\data\\"

def get_origen(nombre):
    if "IBM" in nombre:
        return "IBM"
    elif "TERA" in nombre:
        return "TERA"

def exists_sheet(sheet):
    try:
        wb_result[sheet]
        return True
    except:
        return False

def format_row(row,origen,rownum):
    if origen == "TERA":
        current_row = list(row)
        last_item = current_row[len(current_row) - 1]
        row_formated=current_row[0:len(current_row)-1]+["",last_item]
        if rownum == 1:
            row_formated += ["ORIGEN"]
        else:
            row_formated += [origen]

        return tuple(row_formated)
    else:


for file in listdir(data_dir):
    wb_current = load_workbook(data_dir + file)
    origen = get_origen(file)
    print("archivo:", file, origen)

    #Por cada hoja
    for sheet_name in wb_current.sheetnames:
        print("Hoja:",sheet_name)
        #Crear la hoja nueva en el archivo de resultado
        if not exists_sheet(sheet_name):
            ws_result = wb_result.create_sheet()
            ws_result.title = sheet_name
            nueva = True
        else:
            ws_result = wb_result[sheet_name]
            nueva = False
        
        ws_current = wb_current[sheet_name]
                
        #Copiar el contenido
        contador = 1
        for row in ws_current.values:
            if nueva:
                row_add = format_row(row,origen,contador)
                ws_result.append(row_add)
            else:
                if contador > 1 :
                    row_add = format_row(row,origen,contador)
                    ws_result.append(row_add)
            
            #print(row_add)
            contador +=1

wb_result.save("result.xlsx")
