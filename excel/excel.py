from openpyxl import Workbook
from openpyxl import load_workbook

class Excel:
    def __init__(self):
        pass

    def read_file(self, file_name, sheet_name):
        """Lee un archivo de excel y regresa el contenido en una lita"""
        try:
            wb = load_workbook(filename=file_name)
            ws = wb[sheet_name]
            self.data = [row for row in ws.values]

            return self.data
        except KeyError as key_error:
            print("La hoja {0} no existe".format(sheet_name))
        except FileNotFoundError as fnf_error:
            print("El aarchivo {0} no existe".format(file_name))



    def write_file(self, file_name, sheet_name, data):
        """Escribe el contenido de una lista en un archivo de excel .xslx """
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

        #agregar datos
        for row in data:
            ws.append(row)

        try:
            wb.save(file_name + ".xlsx")
        except:
            print("Ha ocurrido un error al escribir el archivo:",sys.exc_info()[0])

    def get_headers(self):
        if self.data:
            return self.data[0]

    def get_data(self):
        if self.data:
            return self.data[1:]
